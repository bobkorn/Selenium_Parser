from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import time

binary = FirefoxBinary(r"")  # Path to FirefoxBinary
profile = FirefoxProfile(r"")  # Path to FirefoxProfile
pages = range(0, 0, 1)  # Range of pages you want to scrap through

for i in pages:

    driver = webdriver.Firefox(profile, binary)

    driver.get(""+str(i)+"")  # Web-address of each page, where 'i' is its number

    image_path = ""  # Path to image
    status_path = ''  # Path to status

    all_tags = [image_path, status_path]
    finals = {'reg_num': 0, 'req_num': 0, 'exp_date': 0, 'priority': 0,
              'image': 0, 'owner': 0, 'reg_date': 0, 'goods': 0, 'status': 0}

    for tag in all_tags:
        try:
            if tag == image_path:
                text = driver.find_element_by_xpath(tag).get_attribute("href")
                finals['image'] = text
            elif tag == status_path:
                text = driver.find_element_by_xpath(tag).text
                finals['status'] = text
        except:
            continue

    owners = list()
    tags = driver.find_elements_by_class_name("bib")
    tags2 = driver.find_elements_by_class_name("bib2")

    for tag in tags:

        data = tag.text
        if data.startswith(''):
            finals['reg_num'] = data[:]
        elif data.startswith(''):
            finals['req_num'] = data[:]
        elif data.startswith(''):
            finals['exp_date'] = data[:]
        elif data.startswith(''):
            finals['reg_date'] = data[:]
        elif data.startswith(''):
            owners.append(data[:])
        elif data.startswith(''):
            owners.append(data[:])
        elif data.startswith(''):
            finals['goods'] = data[:]

    for tag in tags2:
        finals['priority'] = tag.text

    try:
        finals['owner'] = owners[-1]

    except:
        pass

    driver.quit()
    time.sleep(1)

    df = pd.DataFrame(np.array([[finals['reg_num'], finals['req_num'], finals['exp_date'],
                                 finals['priority'], finals['image'], finals['owner'],
                                 finals['reg_date'], finals['goods'], finals['status']]]),
                       columns=['reg_num', 'req_num', 'exp_date', 'priority',
                            'image', 'owner', 'reg_date', 'goods', 'status'])


    def append_df_to_excel(filename, df, sheet_name='actual_data', startrow=None,
                           truncate_sheet=False, **to_excel_kwargs):

        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            writer.book = load_workbook(filename)

            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            if truncate_sheet and sheet_name in writer.book.sheetnames:

                idx = writer.book.sheetnames.index(sheet_name)

                writer.book.remove(writer.book.worksheets[idx])

                writer.book.create_sheet(sheet_name, idx)


            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            pass

        if startrow is None:
            startrow = 0

        df.to_excel(writer, sheet_name='actual_data', startrow=startrow, index=False, header=None, **to_excel_kwargs)

        writer.save()


    append_df_to_excel('output.xlsx', df)